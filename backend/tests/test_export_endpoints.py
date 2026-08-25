"""新增导出端点测试：异步 PDF 任务 与 Excel 导入模板。

覆盖 create→status→download 全流程、404/409 状态机、错误态，
以及 import-template 的表头与下拉校验（与 scoring_config.yaml 因子对齐）。
"""

import io
import threading
import time

import pytest
from fastapi import FastAPI
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from database import Base, get_db
from models import Customer
from routers import assessment as assessment_router
from routers import customers as customers_router
from services.ai import llm_adapter


class _OfflineAdapter:
    """LLM 离线适配器：保证报告走规则引擎兜底，不触碰外部网络。"""

    available = False
    model = "offline"

    def status(self):
        return {"available": False, "model": self.model}


@pytest.fixture(autouse=True)
def _offline_llm():
    llm_adapter.reset_adapters()
    llm_adapter.set_chat_adapter(_OfflineAdapter())
    yield
    llm_adapter.reset_adapters()


@pytest.fixture()
def engine():
    eng = create_engine(
        "sqlite://", connect_args={"check_same_thread": False}, poolclass=StaticPool
    )
    Base.metadata.create_all(bind=eng)
    try:
        yield eng
    finally:
        Base.metadata.drop_all(bind=eng)


@pytest.fixture()
def session_factory(engine):
    return sessionmaker(bind=engine, autoflush=False, autocommit=False)


@pytest.fixture()
def db(session_factory):
    session = session_factory()
    try:
        yield session
    finally:
        session.close()


@pytest.fixture()
def customer(db):
    c = Customer(customer_name="示例银行(总行)", industry="金融", custom_fields={})
    db.add(c)
    db.commit()
    db.refresh(c)
    return c


@pytest.fixture()
def client(session_factory):
    app = FastAPI()
    app.include_router(assessment_router.router, prefix="/api")
    app.include_router(assessment_router.history_router, prefix="/api")
    app.include_router(customers_router.router, prefix="/api")

    def _override_db():
        session = session_factory()
        try:
            yield session
        finally:
            session.close()

    app.dependency_overrides[get_db] = _override_db
    with TestClient(app) as c:
        yield c


def _wait_ready(client, customer_id, job_id, timeout=30):
    deadline = time.monotonic() + timeout
    while time.monotonic() < deadline:
        resp = client.get(f"/api/assessment/{customer_id}/pdf/jobs/{job_id}")
        assert resp.status_code == 200
        status = resp.json()["status"]
        if status in ("ready", "error"):
            return resp.json()
        time.sleep(0.2)
    raise AssertionError("PDF 任务超时未完成")


def test_pdf_job_create_status_download_flow(client, customer):
    created = client.post(f"/api/assessment/{customer.id}/pdf/jobs")
    assert created.status_code == 200
    job_id = created.json()["job_id"]
    assert created.json()["status"] == "running"

    result = _wait_ready(client, customer.id, job_id)
    assert result["status"] == "ready"
    assert result["error"] is None

    download = client.get(f"/api/assessment/{customer.id}/pdf/jobs/{job_id}/download")
    assert download.status_code == 200
    assert download.content[:4] == b"%PDF"
    assert len(download.content) > 500


def test_pdf_job_download_before_ready_conflicts(client, customer, monkeypatch):
    import routers.assessment as mod

    def _slow_job(*args, **kwargs):
        try:
            time.sleep(3)
        finally:
            mod._PDF_JOB_SEM.release()

    monkeypatch.setattr(mod, "_run_pdf_job", _slow_job)

    created = client.post(f"/api/assessment/{customer.id}/pdf/jobs").json()
    job_id = created["job_id"]

    # 任务仍在 running：下载应返回 409
    download = client.get(f"/api/assessment/{customer.id}/pdf/jobs/{job_id}/download")
    assert download.status_code == 409


def test_pdf_job_unknown_job_returns_404(client, customer):
    resp = client.get(f"/api/assessment/{customer.id}/pdf/jobs/nonexistent")
    assert resp.status_code == 404


def test_pdf_job_customer_mismatch_returns_404(client, customer, db):
    created = client.post(f"/api/assessment/{customer.id}/pdf/jobs").json()
    job_id = created["job_id"]
    other = Customer(customer_name="另一客户", industry="", custom_fields={})
    db.add(other)
    db.commit()

    # 用另一个客户 ID 查询同一个 job：应视为不存在
    resp = client.get(f"/api/assessment/{other.id}/pdf/jobs/{job_id}")
    assert resp.status_code == 404


def test_pdf_job_error_state_reported(client, customer, monkeypatch):
    import routers.assessment as mod

    def _failing_job(job_id, customer_id, include_ai, bind):
        with mod._pdf_jobs_lock:
            job = mod._pdf_jobs.get(job_id)
            if job is not None:
                job["status"] = "error"
                job["error"] = "生成失败：模拟异常"
        mod._PDF_JOB_SEM.release()

    monkeypatch.setattr(mod, "_run_pdf_job", _failing_job)

    created = client.post(f"/api/assessment/{customer.id}/pdf/jobs").json()
    job_id = created["job_id"]

    result = _wait_ready(client, customer.id, job_id)
    assert result["status"] == "error"
    assert "模拟异常" in result["error"]


def test_pdf_job_concurrency_limit_returns_429(client, customer, monkeypatch):
    import routers.assessment as mod

    monkeypatch.setattr(mod, "_PDF_JOB_SEM", threading.BoundedSemaphore(1))
    monkeypatch.setattr(mod, "_run_pdf_job", lambda *args, **kwargs: None)

    first = client.post(f"/api/assessment/{customer.id}/pdf/jobs")
    assert first.status_code == 200

    second = client.post(f"/api/assessment/{customer.id}/pdf/jobs")
    assert second.status_code == 429


def test_pdf_job_sweep_marks_expired_running_as_error(client, customer, monkeypatch):
    """超过 TTL 的 running 标记为 error 并保留（轮询/下载不 404），未过期 running 保留。"""
    import routers.assessment as mod

    monkeypatch.setattr(mod.config, "PDF_JOB_TTL", 60)
    now = time.time()
    with mod._pdf_jobs_lock:
        mod._pdf_jobs["stale"] = {
            "status": "running", "customer_id": customer.id, "created": now - 120,
            "bytes": None, "error": None,
        }
        mod._pdf_jobs["fresh"] = {
            "status": "running", "customer_id": customer.id, "created": now,
            "bytes": None, "error": None,
        }
        mod._sweep_pdf_jobs(now)
        assert mod._pdf_jobs["stale"]["status"] == "error"
        assert "超时" in mod._pdf_jobs["stale"]["error"]
        assert mod._pdf_jobs["fresh"]["status"] == "running"

    # 过期任务仍可查询到 error 状态，而不是 404
    resp = client.get(f"/api/assessment/{customer.id}/pdf/jobs/stale")
    assert resp.status_code == 200
    assert resp.json()["status"] == "error"
    with mod._pdf_jobs_lock:
        mod._pdf_jobs.clear()


def test_pdf_job_cap_evicts_oldest_done_only(client, customer, monkeypatch):
    """超上限时只清最早的 ready/error，running 任务绝不误删。"""
    import routers.assessment as mod

    monkeypatch.setattr(mod, "_PDF_JOBS_MAX", 2)
    now = time.time()
    with mod._pdf_jobs_lock:
        # 清掉其他测试遗留的已完成任务，保证本测试独立可复现
        mod._pdf_jobs.clear()
        mod._pdf_jobs["old_done"] = {
            "status": "ready", "customer_id": customer.id, "created": now - 100,
            "bytes": b"x", "error": None,
        }
        mod._pdf_jobs["new_done"] = {
            "status": "ready", "customer_id": customer.id, "created": now - 50,
            "bytes": b"x", "error": None,
        }
        mod._pdf_jobs["running"] = {
            "status": "running", "customer_id": customer.id, "created": now,
            "bytes": None, "error": None,
        }
        mod._sweep_pdf_jobs(now)
        assert "old_done" not in mod._pdf_jobs
        assert "new_done" in mod._pdf_jobs
        assert "running" in mod._pdf_jobs
        for jid in ("old_done", "new_done", "running"):
            mod._pdf_jobs.pop(jid, None)


def test_import_template_matches_factor_config(client):
    from services.scoring import load_scoring_config

    resp = client.get("/api/customers/import-template")
    assert resp.status_code == 200
    assert resp.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    wb = load_workbook(io.BytesIO(resp.content))
    assert {"使用说明", "因子定义表", "客户数据表", "选项源"} <= set(wb.sheetnames)
    ws = wb["客户数据表"]
    headers = [str(c.value or "").replace("\n", " ") for c in ws[2]]
    assert headers[0] == "客户名称"

    config = load_scoring_config()
    for dim in config.dimensions:
        for f in dim.factors:
            if f.input.type != "readonly":
                assert f.label in headers, f"因子 {f.field} 应出现在导入模板表头"

    # 下拉数据校验已挂载（因子列都允许留空）
    validations = list(ws.data_validations.dataValidation)
    assert len(validations) == 27
    # 下拉引用隐藏的「选项源」工作表，避免内联公式超 255 字符上限
    assert "选项源" in validations[0].formula1
    opt_ws = wb["选项源"]
    assert opt_ws.sheet_state == "hidden"
    assert all(v.allow_blank for v in validations)


def test_v3_xlsx_template_roundtrips_through_import(client):
    template = client.get("/api/customers/import-template")

    # 原样导入：示例行自动跳过并提示，不污染客户库
    resp = client.post(
        "/api/customers/import",
        files={
            "file": (
                "客情因子填报模板_v3.0.xlsx",
                io.BytesIO(template.content),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert resp.status_code == 200
    assert resp.json() == {"created": 0, "errors": ["第3行为模板示例行，已自动跳过"]}

    # 示例行改名为正式客户后再导入：示例值可被导入链路完整解析
    wb = load_workbook(io.BytesIO(template.content))
    ws = wb["客户数据表"]
    ws.cell(row=3, column=1, value="某省政务云客户")
    changed = io.BytesIO()
    wb.save(changed)
    resp2 = client.post(
        "/api/customers/import",
        files={
            "file": (
                "客情因子填报模板_v3.0.xlsx",
                changed.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert resp2.status_code == 200
    assert resp2.json() == {"created": 1, "errors": []}

    customers = client.get("/api/customers", params={"search": "某省政务云客户"}).json()
    imported = customers["items"][0]
    assert imported["custom_fields"]["kcr_01"] == "80-99%"
    assert imported["custom_fields"]["kcr_02"] == "[3,2,2,1,1]"
    assert imported["custom_fields"]["kcr_07"] == "支持≥60%但有反对"
    assert imported["custom_fields"]["risk_08b"] == "中(80-90%)"
    assert imported["custom_fields"]["生命周期阶段"] == "成长"


def test_v3_xlsx_rejects_unrecognizable_values(client):
    template = client.get("/api/customers/import-template")
    wb = load_workbook(io.BytesIO(template.content))
    ws = wb["客户数据表"]
    headers = [str(cell.value or "").replace("\n", " ") for cell in ws[2]]
    kcr_01_column = headers.index("KCR-01 决策链识别完整度") + 1
    ws.cell(row=3, column=kcr_01_column, value="随便填")
    changed = io.BytesIO()
    wb.save(changed)

    resp = client.post(
        "/api/customers/import",
        files={
            "file": (
                "invalid-v3.xlsx",
                changed.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert resp.status_code == 200
    assert resp.json()["created"] == 0
    assert "不在可选项" in resp.json()["errors"][0]


def test_factor_config_includes_rule_text(client):
    """factor-config 下发的 rule_text 应包含可读的打分规则文案。"""
    resp = client.get("/api/customers/factor-config")
    assert resp.status_code == 200
    dims = resp.json()["dimensions"]
    kcr = next(d for d in dims if d["key"] == "kcr")
    kcr_01 = next(f for f in kcr["factors"] if f["field"] == "kcr_01")
    assert "=100% → 10分" in kcr_01["rule_text"]


def test_factor_config_includes_sub_dimension(client):
    """V3.0 已移除二级权重，因子描述保持原子指标口径。"""
    resp = client.get("/api/customers/factor-config")
    assert resp.status_code == 200
    kcr = next(d for d in resp.json()["dimensions"] if d["key"] == "kcr")
    kcr_01 = next(f for f in kcr["factors"] if f["field"] == "kcr_01")
    assert kcr_01["sub_dimension"] == ""
    assert "已识别决策链人数" in kcr_01["description"]
    assert "二级维度" not in kcr_01["description"]
