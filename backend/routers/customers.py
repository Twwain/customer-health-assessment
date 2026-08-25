import datetime
import io
import json
import logging
import re
from typing import Any
from urllib.parse import quote
from fastapi import APIRouter, Depends, HTTPException, Query, Response, UploadFile, File
from sqlalchemy import Boolean, Date, DateTime, Float, Integer, func, or_
from sqlalchemy.orm import Session

from config import SCORING_STRATEGY
from database import get_db
from models import AssessmentHistory, Customer

from schemas import (
    CustomerCreate,
    CustomerUpdate,
    CustomerResponse,
    CustomerListResponse,
    DimensionConfigItem,
    FactorConfigItem,
    FactorConfigResponse,
    FactorInputSpec,
    FactorUpdateRequest,
    FactorUpdateResponse,
    LevelConfigItem,
)
from services import assessment_history
from services.scoring import get_scoring_strategy, load_scoring_config
from services.scoring.config_loader import (
    FactorConfig,
    strip_sub_dimension_annotation,
    sub_dimension_of,
)
import openpyxl

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/customers", tags=["客情信息"])

_RULE_COMPARATORS = {"gte": "≥", "gt": ">", "lte": "≤", "lt": "<", "eq": "="}


def _rule_text(f: FactorConfig) -> str:
    """把打分规则转成前端可读文案（如「≥100% → 2.1分；…；其他 → 0.21分」）。"""
    rule = f.rule
    params = rule.params
    unit = f.input.unit or ""

    def _bracket_text(b) -> str:
        """把单个 bracket 的比较条件拼成文案，支持区间（gte + lt 同时出现）。"""
        parts = []
        for comp in _RULE_COMPARATORS:
            if comp in b:
                parts.append(f"{_RULE_COMPARATORS[comp]}{b[comp]}")
        return "且".join(parts) if parts else ""

    if rule.type == "threshold":
        parts = []
        for b in params.get("brackets", []):
            text = _bracket_text(b)
            if text:
                parts.append(f"{text}{unit} → {b.get('score')}分")
        default = params.get("default") or {}
        if "score" in default:
            parts.append(f"其他 → {default.get('score')}分")
        return "；".join(parts)
    if rule.type == "mapping":
        return "；".join(f"{k} → {v}分" for k, v in (params.get("map") or {}).items())
    if rule.type == "support_distribution":
        return "支持≥60%且无反对 → 10分；有反对者降档 → 8分；支持40-59% → 8分；支持20-39% → 6分；支持<20% → 3分"
    if rule.type == "days_since":
        parts = []
        for b in params.get("brackets", []):
            text = _bracket_text(b)
            if text:
                parts.append(f"{text}天 → {b.get('score')}分")
        empty = params.get("empty") or {}
        if "score" in empty:
            parts.append(f"未填写 → {empty.get('score')}分")
        return "；".join(parts)
    if rule.type == "linear":
        multiplier = params.get("multiplier", 1)
        offset = params.get("offset", 0)
        text = f"得分 = 值 × {multiplier} + {offset}"
        clamp = params.get("clamp") or {}
        if clamp.get("min") is not None or clamp.get("max") is not None:
            text += f"（上限 {clamp.get('max', '—')} / 下限 {clamp.get('min', '—')}）"
        return text
    if rule.type == "penalty":
        return f"条件「{params.get('when', 'truthy')}」命中时 {params.get('score', 0)} 分"
    if rule.type == "constant":
        return f"固定 {params.get('score', 0)} 分"
    return ""


def _build_import_workbook(config) -> tuple[io.BytesIO, list[str]]:
    """构造与 V3.0 填报文件一致的三表 Excel 模板。"""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    factors = []
    for dim in config.dimensions:
        for f in dim.factors:
            if f.input.type == "readonly":
                continue
            factors.append((dim, f))

    base = ["客户名称", "行业", "客户规模", "生命周期阶段"]
    headers = list(base)
    for _, factor in factors:
        code, _, name = factor.label.partition(" ")
        headers.append(f"{code}\n{name or factor.label}")

    wb = Workbook()
    guide = wb.active
    guide.title = "使用说明"
    guide_rows = [
        ["《客情因子填报模板 V3.0》使用说明", ""],
        ["目的", "按精简后的28个因子填写原子指标原始值，由系统自动映射评分"],
        ["操作步骤", "1. 阅读【因子定义表】；2. 在【客户数据表】每个客户填写一行；3. 保持数据自洽并完成脱敏"],
        ["核心原则", "只填原始值，不填0-10分；客户名称、人名一律脱敏"],
    ]
    for row in guide_rows:
        guide.append(row)
    guide.column_dimensions["A"].width = 18
    guide.column_dimensions["B"].width = 90
    guide["A1"].font = Font(bold=True, size=15, color="17365D")
    guide["B1"].font = Font(bold=True, size=15, color="17365D")
    for row in guide.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    definitions = wb.create_sheet("因子定义表")
    definition_headers = [
        "维度", "维度权重", "编号", "因子名称", "原子指标（填什么）",
        "打分映射规则（0-10）", "填报示例", "数据来源", "频率", "填报人",
    ]
    definitions.append(definition_headers)
    for dim, factor in factors:
        code, _, name = factor.label.partition(" ")
        definitions.append([
            dim.name.split(" ", 1)[-1], f"{dim.max_score}%", code, name or factor.label,
            factor.description, _rule_text(factor), factor.example,
            f"{factor.source_role}填报" if factor.source_role != "系统" else "系统自动抓取",
            factor.frequency, "",
        ])

    data_ws = wb.create_sheet("客户数据表")
    hints = ["虚构", "政府/金融/教育/大企业等", "大/中/小", "导入/成长/成熟/衰退"]
    for _, factor in factors:
        source = f"{factor.source_role}填报" if factor.source_role != "系统" else "系统自动抓取"
        hints.append(f"{factor.description}（{source}·{factor.frequency}）")
    data_ws.append(hints)
    data_ws.append(headers)
    data_ws.append([
        "示例-某省政务云客户", "政府", "大", "成长",
        *[_standardize_import_factor_value(factor, factor.example) for _, factor in factors],
    ])

    # 以 = 开头的文本（如规则文案「=100% → 10分」）会被 openpyxl 误判为公式写入 <f>，
    # Excel 打开时报「部分内容有问题」并在修复时删除公式记录；强制按文本存储
    for sheet in (definitions, data_ws):
        for row in sheet.iter_rows():
            for cell in row:
                if cell.data_type == "f":
                    cell.data_type = "s"

    opt_ws = wb.create_sheet("选项源")
    opt_ws.sheet_state = "hidden"
    option_index = 0
    for factor_offset, (_, factor) in enumerate(factors, start=len(base) + 1):
        options = list(factor.input.options)
        # KCR-02 单元格填写五人等级列表，不能套用单值下拉验证。
        if not options or factor.input.type == "key_person_levels":
            continue
        option_index += 1
        for row, opt in enumerate(options, start=1):
            opt_ws.cell(row=row, column=option_index, value=opt)
        dv = DataValidation(
            type="list",
            formula1=(
                f"'选项源'!${get_column_letter(option_index)}$1:"
                f"${get_column_letter(option_index)}${len(options)}"
            ),
            allow_blank=True,
            showErrorMessage=True,
            errorTitle="无效选项",
            error="请从下拉列表中选择有效选项",
        )
        data_ws.add_data_validation(dv)
        letter = get_column_letter(factor_offset)
        dv.add(f"{letter}3:{letter}1048576")

    dark_fill = PatternFill("solid", fgColor="17365D")
    light_fill = PatternFill("solid", fgColor="D9E2F3")
    for sheet, header_row in ((definitions, 1), (data_ws, 2)):
        for cell in sheet[header_row]:
            cell.fill = dark_fill
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for cell in data_ws[1]:
        cell.fill = light_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.font = Font(size=9, color="666666")
    definitions.freeze_panes = "A2"
    data_ws.freeze_panes = "E3"
    for column in range(1, len(headers) + 1):
        letter = get_column_letter(column)
        data_ws.column_dimensions[letter].width = 16 if column <= len(base) else 22
    for column, width in enumerate([18, 12, 12, 25, 42, 42, 42, 16, 10, 12], start=1):
        definitions.column_dimensions[get_column_letter(column)].width = width
    for row in definitions.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf, headers


# 注意：本路由必须声明在 `/{customer_id}` 之前，否则会被当成客户 ID 解析
@router.get("/import-template")
def download_import_template():
    """下载 V3.0 Excel 填报模板。"""
    config = load_scoring_config()
    buf, _ = _build_import_workbook(config)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": (
                'attachment; filename="factor-template-v3.xlsx"; '
                f"filename*=UTF-8''{quote('客情因子填报模板_v3.0.xlsx')}"
            )
        },
    )


@router.post("/import")
def import_customers(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """批量导入 CSV 或 V3.0 Excel 填报文件。"""
    if not file.filename:
        raise HTTPException(status_code=400, detail="请选择文件")

    content = file.file.read()
    filename = file.filename.lower()
    if filename.endswith(".csv"):
        return _import_csv(content, db)
    if filename.endswith(".xls"):
        raise HTTPException(status_code=400, detail="不支持老版 .xls 格式，请另存为 .xlsx")
    if filename.endswith(".xlsx"):
        return _import_excel(content, db)
    raise HTTPException(status_code=400, detail="仅支持 CSV 或 Excel(.xlsx) 文件")


@router.get("", response_model=CustomerListResponse)
def list_customers(
    search: str = Query(default="", description="搜索客户名称/行业/对接人"),
    industry: str = Query(default="", description="按行业筛选"),
    level: str = Query(default="", description="按健康等级筛选（等级名来自 scoring_config.yaml 的 levels）"),
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=20, ge=1, le=1000),
    db: Session = Depends(get_db),
):
    query = db.query(Customer)
    if search:
        query = query.filter(
            or_(
                Customer.customer_name.contains(search),
                Customer.industry.contains(search),
                Customer.contact_person.contains(search),
            )
        )
    if industry:
        query = query.filter(Customer.industry == industry)

    if level:
        # 等级筛选基于最近一次评估快照（AssessmentHistory.level）走 SQL 过滤，
        # 避免对全表逐条实时评分；快照在建档/编辑/因子更新/AI 评估时均会刷新。
        # 无快照的客户（理论上建档即落基线）回退为实时评估，保证语义完整。
        scoring = get_scoring_strategy()
        latest = (
            db.query(
                AssessmentHistory.customer_id,
                func.max(AssessmentHistory.id).label("max_id"),
            )
            .filter(AssessmentHistory.config_version == scoring.config.version)
            .group_by(AssessmentHistory.customer_id)
            .subquery()
        )
        query = query.outerjoin(latest, Customer.id == latest.c.customer_id).outerjoin(
            AssessmentHistory, AssessmentHistory.id == latest.c.max_id
        )
        # 建档/编辑都会写快照；这里只为迁移前的历史客户保留小范围实时回退。
        missing_customers = query.filter(AssessmentHistory.id.is_(None)).all()
        fallback_ids: list[int] = []
        if missing_customers:
            fallback_ids = [c.id for c in missing_customers if scoring.evaluate(c).level == level]
        level_filter = AssessmentHistory.level == level
        if fallback_ids:
            level_filter = or_(level_filter, Customer.id.in_(fallback_ids))
        query = query.filter(level_filter)

    total = query.count()
    items = query.order_by(Customer.updated_at.desc()).offset((page - 1) * page_size).limit(page_size).all()

    return CustomerListResponse(
        items=[CustomerResponse.model_validate(c) for c in items],
        total=total,
        page=page,
        page_size=page_size,
    )


@router.get("/industries")
def list_industries(db: Session = Depends(get_db)):
    results = db.query(Customer.industry).distinct().filter(Customer.industry != "").all()
    return [r[0] for r in results]


# 注意：本路由必须声明在 `/{customer_id}` 之前，否则会被当成客户 ID 解析
@router.get("/factor-config", response_model=FactorConfigResponse)
def get_factor_config():
    """返回当前因子配置（来自 scoring_config.yaml），供前端动态渲染因子表单。

    新增因子只需在配置里注册，前后端均无需改代码。
    """
    config = load_scoring_config()
    return FactorConfigResponse(
        version=config.version,
        updated_at=config.updated_at,
        description=config.description,
        strategy=SCORING_STRATEGY,
        total_max_score=config.total_max_score,
        dimensions=[
            DimensionConfigItem(
                key=dim.key,
                name=dim.name,
                max_score=dim.max_score,
                enabled=dim.enabled,
                description=dim.description,
                factors=[
                    FactorConfigItem(
                        field=f.field,
                        label=f.label,
                        weight=f.weight,
                        source=f.source,
                        source_role=f.source_role,
                        description=strip_sub_dimension_annotation(f.description),
                        sub_dimension=sub_dimension_of(f.description),
                        rule_text=_rule_text(f),
                        rule_type=f.rule.type,
                        editable=f.input.editable,
                        input=FactorInputSpec(
                            type=f.input.type,
                            options=f.input.options,
                            min=f.input.min,
                            max=f.input.max,
                            step=f.input.step,
                            unit=f.input.unit,
                            placeholder=f.input.placeholder,
                        ),
                    )
                    for f in dim.factors
                ],
            )
            for dim in config.dimensions
        ],
        levels=[
            LevelConfigItem(name=lv.name, min_score=lv.min_score, color=lv.color)
            for lv in config.levels
        ],
    )


@router.get("/{customer_id}", response_model=CustomerResponse)
def get_customer(customer_id: int, db: Session = Depends(get_db)):
    customer = db.query(Customer).filter(Customer.id == customer_id).first()
    if not customer:
        raise HTTPException(status_code=404, detail="客户不存在")
    return CustomerResponse.model_validate(customer)


@router.post("", response_model=CustomerResponse)
def create_customer(data: CustomerCreate, db: Session = Depends(get_db)):
    customer = Customer(**data.model_dump())
    db.add(customer)
    db.commit()
    db.refresh(customer)
    # 建档即落一条评估基线，趋势图从第一天就有数据
    assessment_history.record_assessment(
        db, customer, assessed_by="system", trigger="create", skip_if_unchanged=False
    )
    return CustomerResponse.model_validate(customer)


@router.put("/{customer_id}", response_model=CustomerResponse)
def update_customer(customer_id: int, data: CustomerUpdate, db: Session = Depends(get_db)):
    customer = db.query(Customer).filter(Customer.id == customer_id).first()
    if not customer:
        raise HTTPException(status_code=404, detail="客户不存在")
    # 与 PUT /factors 同一校验口径：注册为因子的字段须满足配置中的可选项约束，
    # 防止整体更新绕开因子校验写入非法枚举值，导致规则兜底、分数静默漂移
    config = load_scoring_config()
    for key, value in data.model_dump(exclude_unset=True).items():
        factor = config.find_factor(key)
        if factor is not None and factor.input.options and value not in (None, ""):
            _validate_choice(factor, str(value))
        setattr(customer, key, value)
    db.commit()
    db.refresh(customer)
    # 因子有变化才记历史（仅改备注等非因子字段不会产生新快照）
    assessment_history.record_assessment(db, customer, assessed_by="编辑客户", trigger="update")
    return CustomerResponse.model_validate(customer)


@router.put("/{customer_id}/factors", response_model=FactorUpdateResponse)
def update_customer_factors(
    customer_id: int,
    data: FactorUpdateRequest,
    db: Session = Depends(get_db),
):
    """更新客情因子并即时重算基础客情分。

    - 字段白名单来自 `scoring_config.yaml`，未注册字段会被忽略（返回 ignored_fields）
    - `source: custom_fields` 的扩展因子自动写入 Customer.custom_fields
    - 评审结论 Q6：仅桌面端调用，移动端为只读视图
    """
    customer = db.query(Customer).filter(Customer.id == customer_id).first()
    if not customer:
        raise HTTPException(status_code=404, detail="客户不存在")

    config = load_scoring_config()
    updated_fields: list[str] = []
    ignored_fields: list[str] = []
    custom_fields = dict(customer.custom_fields or {})

    for name, raw_value in data.factors.items():
        factor = config.find_factor(name)
        if factor is None or not factor.input.editable:
            ignored_fields.append(name)
            continue

        if factor.source == "custom_fields":
            custom_fields[name] = _coerce_by_input(factor, raw_value)
        else:
            column = Customer.__table__.columns.get(name)
            if column is None:
                ignored_fields.append(name)
                continue
            setattr(customer, name, _coerce_by_column(factor, column, raw_value))
        updated_fields.append(name)

    if custom_fields != (customer.custom_fields or {}):
        customer.custom_fields = custom_fields

    db.commit()
    db.refresh(customer)

    engine = get_scoring_strategy()
    assessment = engine.evaluate(customer)
    # 因子变更即重算并写入历史，供趋势曲线对比
    assessment_history.record_assessment(
        db, customer, assessment, assessed_by="因子编辑", trigger="factor_update"
    )

    return FactorUpdateResponse(
        customer=CustomerResponse.model_validate(customer),
        assessment=assessment,
        updated_fields=updated_fields,
        ignored_fields=ignored_fields,
    )


def _is_blank(value: Any) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def _validate_choice(factor: FactorConfig, value: Any) -> Any:
    options = factor.input.options
    if options and str(value) not in options:
        raise HTTPException(
            status_code=400,
            detail=f"因子「{factor.label}」的值 `{value}` 不在可选项 {options} 中",
        )
    return value


def _standardize_import_factor_value(factor: FactorConfig, value: Any) -> str:
    """把 V3.0 XLSX 的原子指标值归档到新版标准下拉分档。

    前端/API 仍只接受 options；仅导入流程允许模板中的精确百分比、次数和
    分布列表，并在入库前转换为唯一标准档位，避免把示例值扩张成下拉选项。
    """
    if factor.input.type == "key_person_levels":
        return _parse_key_person_levels(factor, value)

    text = "" if value is None else str(value).strip()
    if not text or text in factor.input.options:
        return text
    if factor.rule.type not in {"threshold", "support_distribution"}:
        _validate_choice(factor, text)

    from services.scoring.rules import evaluate_factor

    if not re.search(r"-?\d", text):
        _validate_choice(factor, text)
    target_score = evaluate_factor(factor, text).score
    candidates = [
        option
        for option in factor.input.options
        if evaluate_factor(factor, option).score == target_score
    ]

    if factor.field == "kcr_07":
        support_match = re.search(r"支持[^\d]*(\d+(?:\.\d+)?)\s*%", text)
        oppose_match = re.search(r"反对[^\d]*(\d+(?:\.\d+)?)\s*%", text)
        support = float(support_match.group(1)) if support_match else -1
        oppose = float(oppose_match.group(1)) if oppose_match else 0
        if support >= 60:
            wanted = "支持≥60%但有反对" if oppose > 0 else "支持≥60%且无反对"
        elif support >= 40:
            wanted = "40-59%支持"
        elif support >= 20:
            wanted = "20-39%支持"
        else:
            wanted = "<20%支持"
        candidates = [option for option in candidates if option == wanted]

    if not candidates:
        _validate_choice(factor, text)
    return candidates[0]


def _validate_range(factor: FactorConfig, number: float) -> float:
    spec = factor.input
    if spec.min is not None and number < float(spec.min):
        raise HTTPException(status_code=400, detail=f"因子「{factor.label}」不能小于 {spec.min}")
    if spec.max is not None and number > float(spec.max):
        raise HTTPException(status_code=400, detail=f"因子「{factor.label}」不能大于 {spec.max}")
    return number


def _parse_number(factor: FactorConfig, value: Any) -> float:
    try:
        return _validate_range(factor, float(value))
    except (TypeError, ValueError):
        raise HTTPException(status_code=400, detail=f"因子「{factor.label}」需要数字，收到 `{value}`")


def _parse_date(factor: FactorConfig, value: Any) -> datetime.date | None:
    if _is_blank(value):
        return None
    if isinstance(value, datetime.datetime):
        return value.date()
    if isinstance(value, datetime.date):
        return value
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y-%m-%dT%H:%M:%S"):
        try:
            return datetime.datetime.strptime(str(value).strip(), fmt).date()
        except ValueError:
            continue
    raise HTTPException(
        status_code=400, detail=f"因子「{factor.label}」日期格式应为 YYYY-MM-DD，收到 `{value}`"
    )


def _parse_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value != 0
    return str(value).strip().lower() in ("1", "true", "yes", "y", "是")


def _parse_key_person_levels(factor: FactorConfig, value: Any) -> str:
    """校验 KCR-02 的五位关键人等级，并存为评分器可解析的 JSON 列表。"""
    raw = value
    if isinstance(value, str):
        text = value.strip()
        try:
            raw = json.loads(text)
        except (json.JSONDecodeError, TypeError):
            raw = [item.strip() for item in text.strip("[]").split(",") if item.strip()]
    if not isinstance(raw, (list, tuple)) or len(raw) != 5:
        raise HTTPException(status_code=400, detail=f"因子「{factor.label}」必须填写 5 位关键人等级")

    levels: list[int] = []
    for item in raw:
        try:
            number = int(item)
        except (TypeError, ValueError):
            raise HTTPException(status_code=400, detail=f"因子「{factor.label}」等级只能为 3、2、1、0、-1")
        if isinstance(item, bool) or str(number) != str(item).strip() or number not in {3, 2, 1, 0, -1}:
            raise HTTPException(status_code=400, detail=f"因子「{factor.label}」等级只能为 3、2、1、0、-1")
        levels.append(number)
    return json.dumps(levels, ensure_ascii=False, separators=(",", ":"))


def _coerce_by_column(factor: FactorConfig, column: Any, value: Any) -> Any:
    """按 Customer 模型列类型做类型转换与校验。"""
    col_type = column.type

    if isinstance(col_type, Boolean):
        return _parse_bool(value)
    if isinstance(col_type, (Date, DateTime)):
        return _parse_date(factor, value)
    if isinstance(col_type, Integer):
        if _is_blank(value):
            return 0
        return int(_parse_number(factor, value))
    if isinstance(col_type, Float):
        if _is_blank(value):
            return 0.0
        return _parse_number(factor, value)

    text = "" if value is None else str(value).strip()
    if text and factor.input.options:
        _validate_choice(factor, text)
    return text


def _coerce_by_input(factor: FactorConfig, value: Any) -> Any:
    """custom_fields 扩展因子按 input.type 做类型转换。"""
    input_type = factor.input.type

    if input_type == "key_person_levels":
        return None if _is_blank(value) else _parse_key_person_levels(factor, value)
    if input_type == "bool":
        return _parse_bool(value)
    if input_type == "date":
        parsed = _parse_date(factor, value)
        return parsed.isoformat() if parsed else None
    if input_type in ("number", "slider"):
        return None if _is_blank(value) else _parse_number(factor, value)
    if _is_blank(value):
        return None

    text = str(value).strip()
    if factor.input.options:
        _validate_choice(factor, text)
    return text


@router.delete("/{customer_id}")
def delete_customer(customer_id: int, db: Session = Depends(get_db)):
    customer = db.query(Customer).filter(Customer.id == customer_id).first()
    if not customer:
        raise HTTPException(status_code=404, detail="客户不存在")
    name = customer.customer_name
    db.delete(customer)
    db.commit()
    logger.info("删除客户 id=%s name=%s（级联清理评估历史与聊天会话）", customer_id, name)
    return {"ok": True}


def _import_csv(content: bytes, db: Session):
    import csv

    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        try:
            # 国内 Excel 导出的 CSV 常见 GBK 编码
            text = content.decode("gbk")
        except UnicodeDecodeError:
            raise HTTPException(status_code=400, detail="CSV 编码无法识别（仅支持 UTF-8 / GBK）")
    reader = csv.DictReader(io.StringIO(text))
    return _process_rows(list(reader), db)


def _import_excel(content: bytes, db: Session):
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Excel 文件无法解析：{exc}")
    ws = wb["客户数据表"] if "客户数据表" in wb.sheetnames else wb.active
    if ws is None:
        raise HTTPException(status_code=400, detail="Excel 中没有可用的工作表")
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="文件为空")

    header_index = next(
        (
            index
            for index, row in enumerate(rows[:10])
            if any(str(value).strip() == "客户名称" for value in row if value is not None)
        ),
        None,
    )
    if header_index is None:
        raise HTTPException(
            status_code=400,
            detail="未找到‘客户名称’表头；V3.0 模板请填写‘客户数据表’工作表",
        )

    headers = [str(h).strip() if h is not None else "" for h in rows[header_index]]
    records = []
    for row in rows[header_index + 1:]:
        if not any(value not in (None, "") for value in row):
            continue
        record = {}
        for i, value in enumerate(row):
            if i < len(headers):
                record[headers[i]] = value
        records.append(record)
    return _process_rows(records, db, start_row=header_index + 2)


def _process_rows(records: list[dict], db: Session, *, start_row: int = 2):
    """批量写入客户行。

    因子列（含 custom_fields 扩展因子）统一走与「编辑因子」相同的
    _coerce_by_column / _coerce_by_input + _validate_choice 校验与类型转换，
    因此枚举值非法会被拒绝、数字/日期/布尔会被正确转换；新增因子只需在
    scoring_config.yaml 注册即可被导入识别，无需改此处代码。
    基础字段（客户名称/行业/对接人/联系电话/备注）走简单字符串映射，保持向后兼容。
    """
    config = load_scoring_config()
    factor_by_field: dict[str, Any] = {}
    factor_by_label: dict[str, Any] = {}
    factor_by_code: dict[str, Any] = {}
    for dim in config.dimensions:
        for f in dim.factors:
            if f.input.type == "readonly":  # 派生/共享因子（如回款逾期扣分）不单独成列
                continue
            # setdefault 保留首次出现：可编辑因子（带 options）优先于同名 readonly 因子
            factor_by_field.setdefault(f.field, f)
            factor_by_label.setdefault(f.label, f)
            code = f.label.split(" ", 1)[0].upper()
            factor_by_code.setdefault(code, f)

    # 基础（非因子）字段：中文表头 -> 模型列
    base_map = {
        "客户名称": "customer_name",
        "行业": "industry",
        "对接人": "contact_person",
        "联系电话": "contact_phone",
        "备注": "notes",
    }
    base_custom_fields = {"客户规模", "生命周期阶段"}
    # 兼容旧模板里与配置 label 不一致的历史中文表头
    legacy_factor_headers = {
        "客户满意度": "customer_satisfaction",
        "合同金额(万元)": "contract_amount",
    }

    created = 0
    errors = []
    imported: list[Customer] = []

    for i, row in enumerate(records):
        try:
            data: dict[str, Any] = {}
            custom_fields: dict[str, Any] = {}

            for cn_key, val in row.items():
                if val is None or val == "":
                    continue
                cn_key = str(cn_key).strip()
                en = base_map.get(cn_key)
                if en:
                    data[en] = str(val).strip()
                    continue
                if cn_key in base_custom_fields:
                    custom_fields[cn_key] = str(val).strip()
                    continue

                # 因子识别：先按配置 label，再按遗留中文表头，最后按英文 field
                code_match = re.match(r"^([A-Za-z]+-\d+[A-Za-z]?)", cn_key)
                factor = factor_by_label.get(cn_key)
                if factor is None and code_match:
                    factor = factor_by_code.get(code_match.group(1).upper())
                if factor is None:
                    factor = factor_by_field.get(legacy_factor_headers.get(cn_key, cn_key))
                if factor is None:
                    factor = factor_by_field.get(cn_key)
                if factor is not None and factor.input.editable:
                    val = _standardize_import_factor_value(factor, val)
                    if factor.source == "custom_fields":
                        custom_fields[factor.field] = _coerce_by_input(factor, val)
                    else:
                        column = Customer.__table__.columns.get(factor.field)
                        if column is None:
                            custom_fields[factor.field] = _coerce_by_input(factor, val)
                        else:
                            data[factor.field] = _coerce_by_column(factor, column, val)
                    continue

                # 其余列归入自定义字段（原始字符串）
                if isinstance(val, (datetime.datetime, datetime.date)):
                    custom_fields[cn_key] = val.isoformat() if isinstance(val, datetime.datetime) else str(val)
                else:
                    custom_fields[cn_key] = str(val).strip()

            if custom_fields:
                data["custom_fields"] = custom_fields

            if "customer_name" not in data:
                errors.append(f"第{start_row + i}行缺少客户名称")
                continue

            customer = Customer(**data)
            db.add(customer)
            imported.append(customer)
            created += 1
        except Exception as e:
            errors.append(f"第{start_row + i}行: {str(e)}")

    db.commit()

    # 导入完成后补一条评估基线，导入的客户也能看到趋势起点
    for customer in imported:
        try:
            assessment_history.record_assessment(
                db, customer, assessed_by="批量导入", trigger="import",
                skip_if_unchanged=False, commit=False,
            )
        except Exception as e:  # 历史记录失败不应影响导入结果
            errors.append(f"{customer.customer_name} 评估基线写入失败: {e}")
    db.commit()

    return {"created": created, "errors": errors}
